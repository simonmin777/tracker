"""
Module description:
using mongoDB to track daily activities
minimal via app:
1. command line as mymoney project
2. -i insert new entry to database
3. -x export entry from database (_id export by default)
4. -u update entry status
5. make 2 module, i.e. class, that works with inmport
"""
# import sys
import re
import openpyxl
# from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from pymongo import MongoClient
# from pymongo.database import Database
# from pymongo.collection import Collection
from pymongo.errors import WriteError
from bson import ObjectId
from typing import List
# from datetime import timedelta
from datetime import datetime

class MongoDB:

    def __init__(self, url, db_name: str, collection_name: str):
        if url is None:
            self.client = MongoClient("mongodb://simon:68paU6OlBbT1FsbA@cluster0-shard-00-00-rbqhf.mongodb.net:27017,cluster0-shard-00-01-rbqhf.mongodb.net:27017,cluster0-shard-00-02-rbqhf.mongodb.net:27017/test?ssl=true&replicaSet=Cluster0-shard-0&authSource=admin&retryWrites=true")
        else:
            self.client = MongoClient(url)
        self.db = self.client.get_database(db_name)
        self.collection = self.db.get_collection(collection_name)

    def __del__(self):
        print('Connection closed')
        self.close_mongodb()

    """
    def connect_mongodb(self, url: str = None):
        if url is None:
            self.client = MongoClient("mongodb://simon:68paU6OlBbT1FsbA@cluster0-shard-00-00-rbqhf.mongodb.net:27017,cluster0-shard-00-01-rbqhf.mongodb.net:27017,cluster0-shard-00-02-rbqhf.mongodb.net:27017/test?ssl=true&replicaSet=Cluster0-shard-0&authSource=admin&retryWrites=true")
        else:
            self.client = MongoClient(url)
    
    def connect_database(cli: MongoClient, name: str) -> Database:
        return cli.get_database(name)
    """
    def close_mongodb(self):
        self.client.close()

    def query_id_exist(self, _id: ObjectId) -> bool:
        cursor = self.collection.find_one({'_id': _id})
        return cursor is not None

    def query_delete(self, _id: ObjectId) -> None:
        self.collection.delete_one({'_id': _id})

    def query_insert_one(self, doc: dict, flag_auto_id) -> bool:
        if flag_auto_id:
            new_doc = {}
            for (key, value) in zip(doc.keys(), doc.values()):
                if key != '_id':
                    new_doc[key] = value
        else:
            new_doc = doc
        try:
            self.collection.insert_one(new_doc)
        except WriteError as error:
            print('Insertion failed with', error.details)
            return False
        return True

    def insert_all(self, doc_list: List[dict], flag_auto_id=True) -> int:
        insert_count = 0
        for doc in doc_list:
            if self.query_insert_one(doc, flag_auto_id):
                insert_count += 1
        return insert_count

    def find_all(self) -> List[dict]:
        cursor = self.collection.find()
        return list(cursor)

    def update_all(self, doc_list: List[dict]) -> int:
        """ find and update using _id by delete and reinster with same _id """
        update_count = 0
        for doc in doc_list:
            if self.query_id_exist(doc['_id']):
                self.query_delete(doc['_id'])
                if self.query_insert_one(doc, flag_auto_id=False):
                    update_count += 1
            else:
                print(doc['_id'], 'not found, skip')
        return update_count

    def find_between_values(self, key: str, start, end) -> List[dict]:
        if start is None and end is None:
            return self.find_all()
        if start is None:
            cursor = self.collection.find({
                key: {'$lt': end}
            })
        elif end is None:
            cursor = self.collection.find({
                key: {'$gte': start}
            })
        else:
            cursor = self.collection.find({
                key: {'$gte': start, '$lt': end}
            })
        return list(cursor)

    def find_key_between_values(self, key1: str, value1, rangekey: str, start, end, flag_eq=True) -> List[dict]:
        if start is None and end is None:
            return []
        operator = '$eq' if flag_eq else '$ne'
        if start is None:
            cursor = self.collection.find({
                rangekey: {'$lt': end},
                key1: {operator: value1},
            })
        elif end is None:
            cursor = self.collection.find({
                rangekey: {'$gte': start},
                key1: {operator: value1},
            })
        else:
            cursor = self.collection.find({
                rangekey: {'$gte': start, '$lt': end},
                key1: {operator: value1}
            })
        return list(cursor)


def _str_to_int(string: str) -> int:
    try:
        rest = int(string)
    except ValueError:
        rest = None
    return rest

def _get_keys(row: tuple) -> dict:
    rest = {}
    for item in row:    # type: Cell
        if item.value is not None and str(item.value) != '':
            rest[item.value] = item.col_idx-1   # col_index start at 1
    return rest

def _get_format_doc_from_row(keys: dict, row: tuple, critical_keys: List = None) -> dict:
    """ return a document with _id field, and there is a date key
    only 5 possible data type: None, str, date, int, IdObject string  """
    doc = {}
    for (key, value) in zip(keys.keys(), keys.values()):
        doc[key] = row[value].value
    # _id field is either None or 24 char string
    if len(str(doc['_id'])) == 24:
        doc['_id'] = ObjectId(doc['_id'])
    else:
        doc['_id'] = None
    # format date field
    datestr = re.sub('[ -]', '', str(doc['date']))
    if isinstance(doc['date'], datetime):
        pass
    elif doc['date'] is None or datestr == '':
        doc['date'] = datetime.today()
    elif datestr.isdigit() and len(datestr) == 8:
        doc['date'] = datetime(int(datestr) // 10000, (int(datestr) // 100) % 100, int(datestr) % 100)
    else:
        doc['date'] = None
    # format status field
    if doc['status'] is None:
        doc['status'] = 'active'
    # check valid entry against critical key list
    if critical_keys is not None and critical_keys != []:
        for key in critical_keys:
            if doc[key] is not None:
                return doc
        doc = None
    return doc

class XlsxFile:

    def __init__(self, filename: str, flag_create=False):
        """ create/open xlsx file, point to sheet index 0 """
        self.filename = filename
        if flag_create:
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(self.filename)
        self.ws = self.wb.worksheets[0]  # type: Worksheet

    def close(self, flag_save=False):
        if flag_save:
            self.wb.save(self.filename)
        self.wb.close()

    def change_sheet(self, index: int = 0, name: str = None) -> bool:
        try:
            if index:
                tmp_ws = self.wb.worksheets[index]
            elif name:
                tmp_ws = self.wb[name]
            else:
                return False
        except IndexError:
            return False
        self.ws = tmp_ws
        return True

    def get_valid_doc_from_xlsx(self, critical_keys: List = None) -> List:
        keys = {}
        doc_list = []
        for i, row in enumerate(self.ws):
            # first row is key string
            if i == 0:
                keys = _get_keys(row)
                continue
            # rest row are data entries, format doc before insertion
            doc = _get_format_doc_from_row(keys, row, critical_keys)
            if doc is not None:
                doc_list.append(doc)
        return doc_list

    def write_doc_to_xlsx(self, doc_list: List[dict]):
        if len(doc_list) == 0:
            return
        # first row is key, write key
        row = 1
        for j, key in enumerate(doc_list[0].keys()):
            self.ws.cell(row, j+1).value = key
        row += 1
        # write content
        for doc in doc_list:
            for j, (key, value) in enumerate(zip(doc.keys(), doc.values())):
                if key == '_id':
                    value = str(value)
                elif key == 'date':
                    value = str(value)[:10]
                self.ws.cell(row, j+1).value = value
            row += 1
        self.wb.save(self.filename)

# end of module

# end of file
